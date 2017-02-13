#!/usr/bin/python
# -*- coding: utf-8 -*-

#
#  This file is part of the `pex100` Python module
#
#  Copyright (c) 2016-2017 - EMBL-EBI
#
#  File author(s): Dénes Türei (turei.denes@gmail.com)
#
#  Distributed under the GPLv3 License.
#  See accompanying file LICENSE.txt or copy at
#      http://www.gnu.org/licenses/gpl-3.0.html
#
#  Website: http://www.ebi.ac.uk/~denes
#

from future.utils import iteritems
from past.builtins import xrange, range, reduce

import os
import sys
import imp
import re
import itertools
import collections
import copy

import xlrd
import openpyxl

import numpy as np
import pandas as pd
import scipy as sp
import scipy.stats as stats
import matplotlib as mpl
import matplotlib.backends.backend_pdf

try:
    import matplotlib_venn as mplvenn
except:
    sys.stdout.write('\t:: Could not import module `matplotlib_venn`.\n')

try:
    import pypath
except:
    sys.stdout.write('\t:: Could not import module `pypath`.\n')

try:
    import pysemsim
except:
    sys.stdout.write('\t:: Could not import module `pysemsim`.\n')

try:
    import kinact
except:
    sys.stdout.write('\t:: Could not import module `kinact`.\n')

if 'unicode' not in __builtins__:
    unicode = str

class Pex100(object):
    
    lFcTableHdr = [
            'uniprot',
            'genesymbol',
            'name',
            'resaa',
            'resnum',
            'label',
            'original_resnum',
            'effect',
            'fc',
            'logfc',
            'zscore',
            'pval',
            'tval'
        ]
    
    def __init__(self,
                 dirBase = '..',
                 fnIdMapping = 'PEX100_Layout.csv',
                 fnXlsFullData = 'AssayData-Cambridge-Peirce-150515.xlsx',
                 fnCombined = 'bmp8_%s.csv',
                 fnFcTable = 'fc_%s_%s.csv',
                 fnFcTopTable = 'fctop_%s_%s.csv',
                 fnFcTopCommon = 'fctop_%s.csv',
                 fnFuncCat = 'func_annot_categories.tab',
                 fnFuncTidy = 'functional.csv',
                 fnKinactTop = 'kinact_top.tab',
                 fnVennPlot = 'FC_Venn_%s_%s.pdf',
                 ncbi_tax_id = 10090,
                 org_strict = True,
                 flex_resnum = True):
        """
        This is the main class of the module representing an analysis
        on a set of phosphorylation sites across multiple samples.
        It provides various methods and workflows for its analysis.
        
        :param str dirBase: The base directory.
        :param str fnIdMapping: The ID mapping (layout) file of the assay.
                                This is necessary because PEX100 assay
                                uses non-standard names by default.
        :param str fnXlsFullData: The phosphoassay data in xls file.
        :param int ncbi_tax_id: The NCBI Taxonomy ID of the species to
                                work with. Not necessary to provide your
                                experimental organism here, it is safe to
                                let it be human even if the data is from mice.
    
        """
        
        self.dirBase = dirBase
        
        self.set_path(fnXlsFullData, 'fnXlsFullData')
        self.set_path(fnIdMapping, 'fnIdMapping')
        self.set_path(fnCombined, 'fnCombined')
        self.set_path(fnFcTable, 'fnFcTable')
        self.set_path(fnFcTopTable, 'fnFcTopTable')
        self.set_path(fnFcTopCommon, 'fnFcTopCommon')
        self.set_path(fnFuncCat, 'fnFuncCat')
        self.set_path(fnFuncTidy, 'fnFuncTidy')
        self.set_path(fnKinactTop, 'fnKinactTop')
        self.set_path(fnVennPlot, 'fnVennPlot')
        
        self.reAnnot = re.compile(r'([\-\s/\.,\(\)\+A-Za-z0-9]{2,}) '\
            r'\(([A-Z][a-z]+)-?([A-Za-z0-9/]*)\)')
        
        self.reRes  = re.compile(r'([A-Z]?[a-z]*)([0-9]+)')
        self.reORes = re.compile(r'[A-Za-z]{0,3}([/0-9]+)+')
        self.rePsite = re.compile(r'([A-Z0-9]+)_([A-Z])([0-9]+)')
        
        self.dAaletters = {
            'Thr': 'T',
            'Tyr': 'Y',
            'Ser': 'S'
        }
        self.dOrg = {
            'H': 9606,
            'M': 10090,
            'R': 10116
        }
        
        self.ncbi_tax_id = ncbi_tax_id
        self.org_strict = org_strict
        self.flex_resnum = flex_resnum
        
        self.standards = ['Beta actin', 'GAPDH', 'PKC pan activation site']
        self.dStd = {}
        
        self.lDataCols = [('Control', 0), ('BMP8b', 1), ('NE', 2), ('BMP8b_NE', 3)]
        
        self.dMissingHomologs = {
            'IKK-alpha': 'Q60680',
            'IKK-alpha/beta': 'Q60680',
            'GAPDH': 'P16858',
            'Calmodulin': 'P62204',
            'TYK2': 'Q9R117',
            'Beta actin': 'P60710',
            'Met': 'P16056',
            'ATP1A1/Na+K+ ATPase1': 'Q8VDN2',
            'GRK1': 'Q9WVL4',
            'MAP3K7/TAK1': 'Q62073',
            'Rel': 'P15307',
            'Gab2': 'Q9Z1S8',
            'DDX5/DEAD-box protein 5': 'Q61656',
            'MARCKS': 'P26645',
            'PLD1': 'Q9Z280',
            'CD45': 'P06800',
            'Pim-1': 'P06803',
            'MAP3K1/MEKK1': 'P53349',
            'Progesterone Receptor': 'Q00175',
            'PAK1': 'O88643',
            'RapGEF1': 'Q8C5V7',
            'PAK1/2': ['O88643', 'Q8CIN4'],
            'PAK1/2/3': ['O88643', 'Q8CIN4', 'Q61036'],
            'CD227/mucin 1': 'Q02496',
            'HDAC3': 'O88895',
            'TSC2': 'Q61037',
            'Tuberin/TSC2': 'Q61037',
            'CaMK4': 'P08414',
            'STAT2': 'Q9WVL2',
            'AurB/C': ['O70126', 'O88445'],
            'JAK1': 'P52332',
            'NFAT4': 'P97305',
            'HDAC5': 'Q9Z2V6',
            'VEGFR2': 'P35918',
            'STAT1': 'P42225',
            'EPHA2/3/4': ['Q03145', 'Q03137', 'P29319'],
            'Histone H3.1': ['P68433', 'P84228', 'P84244']
        }
        
        self.dAssayRefCorrection = {
            ('Kv1.3/KCNA3', 135): {
                9606: ('Y', 187),
                10090: ('Y', 140),
                10116: ('Y', 137)
            },
            ('PKC delta', 505): {
                9606: ('T', 507)
            },
            ('Opioid Receptor', 375): {
                9606: ('S', 377)
            },
            ('IkB-epsilon', 22): {
                9606: ('S', 161)
            },
            ('HSL', 563): {
                9606: ('S', 853),
                10090: ('S', 557),
                10116: ('S', 863)
            },
            ('KSR', 392): {
                9606: ('S', 406),
                10116: ('S', 40)
            },
            ('BAD', 155): {
                9606: ('S', 118),
                10116: ('S', 156)
            },
            ('BAD', 112): {
                9606: ('S', 75),
                10116: ('S', 113)
            },
            # this EPHB2
            ('EPHB1/2', 604): {
                9606: ('Y', 596),
                10090: ('Y', 596),
                10116: ('Y', 709)
            },
            ('NMDAR2B', 1472): {
                9606: ('Y', 1474)
            },
            ('SRF', 99): {
                9606: ('S', 103)
            },
            ('JAK1', 1022): {
                9606: ('Y', 1034),
                10090: ('Y', 1033)
            },
            ('IkB-beta', 19): {
                9606: ('S', 19)
            },
            ('Myosin regulatory light chain 2', 18): {
                9606: ('S', 20),
                10090: ('S', 20),
                10116: ('S', 20)
            },
            ('BAD', 128): {
                9606: ('S', 91),
                10116: ('S', 129)
            },
            ('BAD', 136): {
                9606: ('S', 99),
                10116: ('S', 137)
            },
            ('RelB', 552): {
                9606: ('S', 573)
            },
            ('IL3RB', 593): {
                10090: ('Y', 595)
            },
            ('BIM', 65): {
                9606: ('S', 69)
            },
            ('CaMK4', 196): {
                9606: ('T', 200)
            }
        }
    
    def reload(self):
        modname = self.__class__.__module__
        mod = __import__(modname, fromlist = [modname.split('.')[0]])
        imp.reload(mod)
        new = getattr(mod, self.__class__.__name__)
        setattr(self, '__class__', new)
    
    #
    # Loading experimental data.
    #
    
    def main(self, network = False):
        """
        Loads and preprocesses all input data.
        Runs the whole data processing and analysis and exports files.
        """
        
        self.init_pypath()
        self.idmapping()
        self.read_tables()
        self.load_ptms()
        self.load_seq()
        self.create_ptms()
        self.ptms_lookup()
        if network:
            self.network()
        self.combined_table(to_file = True)
        self.fc_table(to_file = False)
        self.regulatory_sites()
        self.fc_table(to_file = True)
        self.load_go()
        self.fc_top_table()
        self.kinact_analysis()
        self.kinact_top(threshold = 1.0, fname = 'kinact_top.tab')
        self.functional_array()
        self.tidy_functional_array()
        sys.stdout.write('\t:: Tests without effect sign:\n')
        self.functional_compare_fcs()
        sys.stdout.write('\t:: Tests with effect sign:\n')
        self.functional_compare_fcs(signs = True)
    
    def network(self, extra_proteins = [], edges_percentile = 50, pfile = None):
        """
        Loads and processes the network.
        """
        self.collect_proteins()
        self.load_network(pfile = pfile)
        self.get_network(keep_also = extra_proteins)
        self.sparsen_network(perc = edges_percentile)
    
    def set_path(self, fname, attr):
        """
        Sets the path to a file in basedir.
        """
        setattr(self, attr, os.path.join(self.dirBase, fname))
    
    def init_pypath(self):
        """
        Initializes a PyPath object for mapping and network building.
        """
        self.pa = pypath.PyPath(ncbi_tax_id = self.ncbi_tax_id)
    
    def read_tables(self):
        """
        Reads all data tables.
        """
        self.read_signals()
        self.read_coeffvar()
        self.read_normalized()
        self.read_psites_diffs()
    
    def idmapping(self):
        """
        Builds the idmapping table in order to translate the custom names
        to UniProts already at loading the table.
        """
        self.read_idmapping()
        self.read_organism_specificities()
        if self.ncbi_tax_id == 10090:
            self.mousedict()
            self.idmapping2mouse()
            self.dNamesUniprots = self.dNamesMouseUniprots
        else:
            self.dNamesHumanUniprots = \
                dict(
                    map(
                        lambda x:
                            (
                                x[0],
                                x[1][0]
                            ),
                        iteritems(self.dNamesIds)
                    )
                )
            
            self.dNamesUniprots = self.dNamesHumanUniprots
        
        self.dOrgSpecUniprot = \
            dict(
                itertools.chain(
                    *map(
                        lambda i:
                            map(
                                lambda u:
                                    (
                                        (u, i[0][1], i[0][2]),
                                        i[1]
                                    ),
                                self.dNamesUniprots[i[0][0]]
                            ),
                        iteritems(self.dOrgSpec)
                    )
                )
            )
    
    def read_psites_diffs(self):
        """
        Reads the measured values for each phosphosite at each condition.
        """
        
        self.read_data('Psite', 16, 9, 7, 582, 3)
    
    def read_signals(self):
        """
        Reads the mean of duplicates signal values.
        """
        
        self.read_data('Signal', 0, 9, 5, 1320, 1)
    
    def read_coeffvar(self):
        """
        Reads coefficient of variation values for each pair of spots.
        """
        
        self.read_data('CVar', 0, 9, 10, 1320, 6)
    
    def read_normalized(self):
        """
        Reads the values normalized to median.
        """
        
        self.read_data('Norm', 0, 9, 15, 1320, 11)
    
    def read_data(self, name, col, row, ncol, nrow, vcol):
        """
        Reads a data table with its annotations.
        """
        
        if not hasattr(self, 'llFullData') or self.llFullData is None:
            self.read_raw()
        
        name = name.capitalize()
        
        attr = 'll%s' % name
        
        setattr(self, attr,
                self.ll_table_slice(self.llFullData, col, row, ncol, nrow))
        
        annot, data = self.get_arrays(getattr(self, attr), vcol)
        
        setattr(self, 'a%sAnnot' % name, annot)
        setattr(self, 'a%sData' % name, data)
        
        self.dStd[name] = self.get_standards(name, vcol)
    
    def get_standards(self, name, vcol):
        """
        Gets the values from the standards.
        """
        return \
            dict(
                map(
                    lambda l:
                        (
                            l[0],
                            list(map(float, l[vcol:]))
                        ),
                    filter(
                        lambda l:
                            l[0] in self.standards,
                        getattr(self, 'll%s' % name)
                    )
                )
            )
    
    def read_raw(self):
        """
        Reads the full xls table.
        """
        self.llFullData = self.read_xls(self.fnXlsFullData)
    
    def read_xls(self, xls_file, sheet = 0, csv_file = None,
        return_table = True):
        """
        Generic function to read MS Excel XLS file. Converts one sheet
        to CSV, or returns as a list of lists
        """
        table = []
        try:
            book = xlrd.open_workbook(xls_file, on_demand = True)
            try:
                if type(sheet) is int:
                    sheet = book.sheet_by_index(sheet)
                else:
                    sheet = book.sheet_by_name(sheet)
            except xlrd.biffh.XLRDError:
                sheet = book.sheet_by_index(0)
            table = [[unicode(c.value) \
                for c in sheet.row(i)] \
                for i in xrange(sheet.nrows)]
        except IOError:
            sys.stdout.write('No such file: %s\n' % xls_file)
            sys.stdout.flush()
        except:
            try:
                book = openpyxl.load_workbook(filename = xls_file,
                    read_only = True)
            except:
                sys.stdout.write('\tCould not open xls: %s\n' % xls_file)
                if not os.path.exists(xls_file):
                    sys.stdout.write('\tFile does not exist.\n')
                sys.stdout.flush()
            try:
                if type(sheet) is int:
                    sheet = book.worksheets[sheet]
                else:
                    sheet = book[sheet]
            except:
                sheet = book.worksheets[0]
            cells = sheet.get_squared_range(1, 1,
                sheet.max_column, sheet.max_row)
            table = \
                list(
                    map(lambda row:
                        list(
                            map(lambda c:
                                unicode(c.value),
                                row
                            )
                        ),
                        cells
                    )
                )
        if csv_file:
            with open(csv_file, 'w') as csv:
                csv.write('\n'.join(['\t'.join(r) for r in table]))
        if not return_table:
            table = None
        if 'book' in locals() and hasattr(book, 'release_resources'):
            book.release_resources()
        return table
    
    def ll_table_slice(self, table, col, row, ncol = None, nrow = None):
        """
        Returns a slice from a list-of-lists table, starting at (row, col),
        height and width of (nrow, ncol).
        """
        cend = None if ncol is None else col + ncol
        rend = None if nrow is None else row + nrow
        return list(map(lambda l: list(l)[col:cend], table[row:rend]))
    
    def get_arrays(self, llTable, vcol):
        """
        Returns annotations and data as numpy arrays.
        Requires list-of-lists table and
        the column number where data values start.
        """
        
        return \
            list(
                map(
                    lambda a:
                        np.array(a[0], dtype = a[1]),
                    # zip data and dtype
                    zip(
                        # zip annotations and values
                        zip(
                            *list(
                                # filter organism mismatches
                                filter(
                                    lambda r:
                                        not self.org_strict or \
                                            self.ncbi_tax_id in \
                                            self.dOrgSpec[(r[0][1], r[0][3], r[0][4])],
                                    # chain rows & residues
                                    itertools.chain(
                                        *map(
                                            # iterate rows
                                            # here one row is one antibody
                                            # with enumerate we get an
                                            # unique ID for each antibody
                                            lambda ill:
                                                list(
                                                    itertools.chain(
                                                        *map(
                                                            # iterate residues
                                                            lambda res:
                                                                map(
                                                                    lambda u:
                                                                        (
                                                                            # annotation
                                                                            [
                                                                                u,
                                                                                ill[1][0][0],
                                                                                ill[1][0][1],
                                                                                res[0],
                                                                                res[1],
                                                                                ill[0], # the antibody ID
                                                                                ''.join(self.reORes.findall(ill[1][0][2]))
                                                                                # the original residue numbers:
                                                                                # sometimes these are needed for
                                                                                # an unambiguous identification
                                                                                # of the antibodies
                                                                            ],
                                                                            # values
                                                                            ill[1][1]
                                                                        ),
                                                                    # one name might be mapped to
                                                                    # multiple UniProt IDs
                                                                    self.dNamesUniprots[ill[1][0][0]]
                                                                ),
                                                            # one antibody might detect multiple
                                                            # phosphorylations
                                                            self.get_residues(ill[1][0][2])
                                                        )
                                                    )
                                                ),
                                            enumerate(
                                                map(
                                                    # for each row, match annotation
                                                    # and convert data to float
                                                    lambda l:
                                                        (
                                                            self.reAnnot.\
                                                                match(l[0]).groups(0),
                                                            list(map(float, l[vcol:]))
                                                        ),
                                                    # starting from list of lists
                                                    filter(
                                                        # filter standards
                                                        lambda l:
                                                            l[0] not in self.standards,
                                                        llTable
                                                    )
                                                )
                                            )
                                        )
                                    )
                                )
                            )
                        ),
                        # dtypes for annot. and data
                        [np.object, np.float]
                    )
                )
            )
    
    def get_residues(self, res):
        """
        Matches residues in string,
        returns tuples of tuples with residue names and numbers.
        """
        ress = self.reRes.findall(res)
        aa = ''
        sres = []
        for r in ress:
            if r[0] in self.dAaletters:
                aa = self.dAaletters[r[0]]
            sres.append((aa, int(r[1])))
        return tuple(sres)
    
    def read_organism_specificities(self):
        """
        Reads the organism specificities of the assay probes.
        These are either for human, mouse, rat, or any combination of these.
        """
        with open(self.fnIdMapping, 'r') as f:
            self.dOrgSpec = \
                dict(
                    itertools.chain(
                        *map(
                            lambda p:
                                map(
                                    lambda res:
                                        (
                                            (p[0][0], res[0], res[1]),
                                            list(
                                                map(
                                                    lambda s:
                                                        self.dOrg[s],
                                                    p[1]
                                                )
                                            )
                                        ),
                                    self.get_residues(p[0][2])
                                ),
                            map(
                                lambda l:
                                    (
                                        self.reAnnot.match(l[1]).groups(0),
                                        l[2]
                                    ),
                                filter(
                                    lambda l:
                                        l[1].strip() not in self.standards,
                                    map(
                                        lambda l:
                                            l.split('\t'),
                                        f
                                    )
                                )
                            )
                        )
                    )
                )
    
    def read_idmapping(self):
        """
        Reads the mapping between the assay custom names and
        standard IDs (UniProt and Entrez Gene).
        """
        with open(self.fnIdMapping, 'r') as f:
            self.dNamesIds = \
                dict(
                    map(
                        lambda l:
                            (
                                self.reAnnot.match(l[1]).groups(0)[0] \
                                    if l[1] not in self.standards \
                                    else l[1], # name
                                (
                                    list(
                                        map(
                                            lambda u:
                                                u.strip(),
                                            l[3].strip().split('/'), # uniprot
                                        )
                                    ),
                                    list(
                                        map(
                                            lambda e:
                                                e.strip(),
                                            l[4].strip().replace(
                                                '\xa0', '').split('/')# entrez
                                        )
                                    )
                                )
                            ),
                        map(
                            lambda l:
                                l.split('\t'),
                            f
                        )
                    )
                )
    
    def idmapping2mouse(self):
        """
        Creates a dict of names to mouse UniProt IDs.
        """
        self.dNamesMouseUniprots = \
            dict(
                map(
                    lambda i:
                        (
                            i[0],
                            list(
                                itertools.chain(
                                    *map(
                                        lambda e:
                                            self.pa.mapper.map_name(e,
                                                                    'entrez',
                                                                    'uniprot',
                                                                    10090),
                                        reduce(
                                            lambda s1, s2:
                                                s1 | s2,
                                            map(
                                                lambda h:
                                                    self.dHuman2Mouse[h],
                                                filter(
                                                    lambda h:
                                                        h in self.dHuman2Mouse,
                                                    i[1][1]
                                                )
                                            ),
                                            set([])
                                        )
                                    )
                                )
                            )
                        ),
                    iteritems(self.dNamesIds)
                )
            )
        
        # adding missing homologs
        for n in self.dNamesMouseUniprots.keys():
            if n in self.dMissingHomologs:
                m = self.dMissingHomologs[n]
                if type(m) is list:
                    self.dNamesMouseUniprots[n].extend(m)
                else:
                    self.dNamesMouseUniprots[n].append(m)
                self.dNamesMouseUniprots[n] = \
                    list(set(self.dNamesMouseUniprots[n]))
    
    def mousedict(self):
        """
        Reads human to mouse mapping from HomoloGene.
        """
        self.dHuman2Mouse = pypath.dataio.homologene_dict(9606, 10090, 'entrez')
    
    #
    # Loading database data.
    #
    
    def load_ptms(self):
        """
        Obtains a list of all kinase-substrate interactions from
        Signor, phosphoELM, PhosphoSite and dbPTM.
        Creates a dict `dPhosDb` where (kinase, substrate) UniProts are keys
        and list of PTMs are values.
        """
        self.dPhosDb = {}
        
        if self.ncbi_tax_id == 10090:
            lKinSub = itertools.chain(
                self.pa.load_signor_ptms(return_raw=True),
                self.pa.load_phosphoelm(return_raw=True),
                self.pa.load_dbptm(return_raw=True),
                self.pa.load_psite_phos(return_raw=True)
            )
        else:
            lKinSub = itertools.chain(
                self.pa.load_signor_ptms(return_raw=True),
                self.pa.load_li2012_ptms(return_raw=True),
                self.pa.load_hprd_ptms(return_raw=True),
                self.pa.load_mimp_dmi(return_raw=True),
                self.pa.load_pnetworks_dmi(return_raw=True),
                self.pa.load_phosphoelm(return_raw=True),
                self.pa.load_dbptm(return_raw=True),
                self.pa.load_psite_phos(return_raw=True)
            )
        
        for ksub in lKinSub:
            if ksub.ptm.protein not in self.dPhosDb:
                self.dPhosDb[ksub.ptm.protein] = []
            self.dPhosDb[ksub.ptm.protein].append(ksub)
    
    def load_seq(self):
        """
        Loads protein sequences from UniProt.
        """
        self.dSeq = pypath.uniprot_input.swissprot_seq(self.ncbi_tax_id,
                                                       isoforms = True)
    
    def create_ptms(self):
        """
        Creates PTM objects for each PTM on array.
        """
        self.dAssaySub = {}
        self.setSeqMismatch = set([])
        self.setSeqMismatchName = set(map(lambda p: (p[1], p[3], p[4]),
                                          self.aPsiteAnnot))
        self.setSeqMismatchOrg = set([])
        self.setSeqMismatchNameOrg = set(
                                      filter(
                                          lambda k:
                                              self.ncbi_tax_id in self.dOrgSpec[k],
                                            map(lambda p: (p[1], p[3], p[4]),
                                                self.aPsiteAnnot)
                                        ))
        
        self.dResUpdate = {}
        
        for psite in self.aPsiteAnnot:
            
            key = (psite[0], psite[3], psite[4])
            nkey = (psite[1], psite[3], psite[4])
            
            # 
            self.setSeqMismatch.add(key)
            
            if psite[0] in self.dSeq:
                
                for isof in self.dSeq[psite[0]].isoforms():
                    
                    resnum = psite[4]
                    resaa = psite[3]
                    
                    if (psite[1], resnum) in self.dAssayRefCorrection and \
                        self.ncbi_tax_id in self.dAssayRefCorrection[
                        (psite[1], resnum)]:
                        
                        resaa, resnum = \
                            self.dAssayRefCorrection[(psite[1], resnum)][
                                self.ncbi_tax_id]
                    
                    if not self.dSeq[psite[0]].match(resaa,
                                                     resnum, isoform = isof):
                        resnum = psite[4] + 1
                        if not self.flex_resnum or \
                            not self.dSeq[psite[0]].match(resaa,
                                                     resnum, isoform = isof):
                            continue
                        
                    if self.dSeq[psite[0]].match(resaa,
                                                     resnum, isoform = isof):
                        
                        # updating the residue number in array
                        if resnum != psite[4]:
                            self.dResUpdate[(psite[0], psite[4])] = \
                                (psite[0], resnum)
                            psite[4] = resnum
                        
                        res = pypath.intera.Residue(resnum,
                                                    resaa,
                                                    psite[0],
                                                    isoform = isof)
                        reg = self.dSeq[psite[0]].get_region(
                            residue = resnum, flanking = 7, isoform = isof)
                        mot = pypath.intera.Motif(psite[0], reg[0], reg[1],
                                                  instance = reg[2],
                                                  isoform = isof)
                        ptm = pypath.intera.Ptm(psite[0],
                                                residue = res,
                                                motif = mot,
                                                typ = 'phosphorylation',
                                                isoform = isof)
                        
                        if psite[0] not in self.dAssaySub:
                            self.dAssaySub[psite[0]] = []
                        self.dAssaySub[psite[0]].append(ptm)
                        
                        # this matched once, so removed to
                        self.setSeqMismatch.remove(key)
                        if nkey in self.setSeqMismatchName:
                            self.setSeqMismatchName.remove(nkey)
                        if nkey in self.setSeqMismatchNameOrg:
                            self.setSeqMismatchNameOrg.remove(nkey)
                        
                        break # process only the first matching isoform
            
            else:
                sys.stdout.write('\t!! No sequence available for %s!\n' % \
                    psite[0])
            
            if key in self.setSeqMismatch:
                
                # here the organism matches, but the sequence mismatches:
                if self.ncbi_tax_id in self.dOrgSpec[nkey]:
                    self.setSeqMismatchOrg.add(key)
            
        # updating residue numbers in all arrays:
        # arrays
        for a in [self.aSignalAnnot, self.aCvarAnnot, self.aNormAnnot]:
            # rows
            for r in a:
                key = (r[0], r[4])
                if key in self.dResUpdate:
                    r[4] = self.dResUpdate[key][1]
        
        for k in list(self.dOrgSpecUniprot.keys()):
            key = (k[0], k[2])
            if key in self.dResUpdate:
                self.dOrgSpecUniprot[(k[0], k[1], self.dResUpdate[key][1])] = \
                    self.dOrgSpecUniprot[k]
                del self.dOrgSpecUniprot[k]
    
    def combined_table(self, to_file = False):
        """
        Export a table with number of kinases for each substrate PTM.
        The table will be saved into `daTable` attribute, which is a
        dict of arrays.
        
        :param bool to_file: Whether to write table into file.
        """
        
        def get_pratio(a, ckey, pkey, lnums, cnum):
            return a[lnums[pkey],cnum] / a[lnums[ckey],cnum]
        
        self.lCombinedHdr = [
                          'uniprot', 'gsymbol', 'name',
                          'numof_kin', 'degree',
                          'resaa', 'resnum', 'group', 'label',
                          'original_resnum',
                          'std_gapdh', 'std_actin', 'std_pkc',
                          'signal', 'ctrl_signal',
                          'cv', 'ctrl_cv',
                          'norm',
                          'norm_actin',
                          'norm_gapdh',
                          'norm_pkc',
                          'ctrl_norm',
                          'ctrl_norm_actin',
                          'ctrl_norm_gapdh',
                          'ctrl_norm_pkc',
                          'phos',
                          'pratio', 'ctrl_pratio',
                          'pratio_actin', 'ctrl_pratio_actin',
                          'pratio_gapdh', 'ctrl_pratio_gapdh',
                          'pratio_pkc', 'ctrl_pratio_pkc',
                          'fc', 'fc_actin', 'fc_gapdh', 'fc_pkc'
                          ]
        
        self.lSingleHdr = [
                          'uniprot', 'gsymbol', 'name',
                          'numof_kin', 'degree',
                          'resaa', 'resnum', 'group', 'label',
                          'original_resnum',
                          'phos',
                          'cv', 'ctrl_cv',
                          'std_name',
                          'std_signal',
                          'norm',
                          'ctrl',
                          'sd_treat',
                          'sd_ctrl',
                          'pratio',
                          'ctrl_pratio',
                          'fc'
                          ]
        
        # initializing data structures
        ctrls = {
            'actin': 'Beta actin',
            'gapdh': 'GAPDH',
            'pkc':   'PKC pan activation site'
        }
        
        dllCombined = {'none': [], 'actin': [], 'gapdh': [], 'pkc': []}
        
        self.daSignalNorm = {}
        
        self.daStd = {}
        
        dDataLnum = dict(map(lambda i:
                                 (
                                    (i[1][0], i[1][2],
                                    i[1][3], i[1][4],
                                    i[1][6] # the original residue numbers
                                    ),
                                    i[0] # the row number
                                 ),
                             enumerate(self.aSignalAnnot)))
        self.dDataLnum = dDataLnum
        
        # standards values normalized by their min
        for stdshort in ctrls.keys():
            
            self.daStd[stdshort] = np.array(self.dStd['Signal'][ctrls[stdshort]])
            self.daStd[stdshort] = (self.daStd[stdshort] /
                                    self.daStd[stdshort].min())
        
        # normalized to standard
        for stdshort in ctrls.keys():
            
            #self.daSignalNorm[stdshort] = self.aSignalData / self.daStd[stdshort]
            # normalized to median
            self.daSignalNorm[stdshort] = self.aSignalData / np.median(self.aSignalData, axis = 0)
            self.daSignalNorm[stdshort] = self.daSignalNorm[stdshort] / self.daStd[stdshort]
            #self.daSignalNorm[stdshort] = (self.daSignalNorm[stdshort] /
            #                               np.median(self.daSignalNorm[stdshort],
            #                                         axis = 0))
        
        # degree of freedom
        self.iDf = self.aPsiteAnnot.shape[0] - 1
        
        self.cntrUniqueLabels = collections.Counter()
        
        for i, annot in enumerate(self.aPsiteAnnot):
            
            uniprot = annot[0]
            gss = self.pa.mapper.map_name(uniprot, 'uniprot',
                                          'genesymbol', self.ncbi_tax_id)
            genesymbol = gss[0] if len(gss) else uniprot
            key = (uniprot, annot[3], annot[4])
            pkey = (uniprot, 'Phospho', annot[3], annot[4], annot[6])
            ckey = (uniprot, 'Ab',      '',       annot[4], annot[6])
            degree = self.pa.graph.vs.select(name = uniprot)[0].degree() \
                if uniprot in self.pa.graph.vs['name'] else 0
            
            if key not in self.dKinNum:
                # this verifies that the phosphorylated residue
                # has been found in the UniProt sequence
                continue
            
            dCtrlPratio = {'none': get_pratio(self.aSignalData,
                                              ckey, pkey, dDataLnum, 0)}
            
            # get the unique label
            label = self.antibody_id_to_name(annot[5])
            
            # phosphorylation ratios for each standards
            for stdshort in ctrls.keys():
                
                dCtrlPratio[stdshort] = \
                    get_pratio(self.daSignalNorm[stdshort],
                               ckey, pkey, dDataLnum, 0)
            
            for group, cnum in self.lDataCols:
                
                dPratio = {'none': get_pratio(self.aNormData,
                                              ckey, pkey,
                                              dDataLnum, cnum)}
                
                # phosphorylation ratios for each standards
                for stdshort in ctrls.keys():
                    
                    dPratio[stdshort] = \
                        get_pratio(self.daSignalNorm[stdshort],
                                ckey, pkey, dDataLnum, cnum)
                
                for dkey, phos in [(ckey, 'np'), (pkey, 'p')]:
                    
                    common_fields = [
                        uniprot,
                        genesymbol,
                        annot[1],
                        self.dKinNum[key],
                        degree,
                        annot[3],
                        annot[4],
                        group,
                        label,
                        annot[6],
                        phos,
                        self.aCvarData[dDataLnum[dkey],cnum], # CV treatment
                        self.aCvarData[dDataLnum[dkey],0]     # CV control
                        # 10 (index)
                    ]
                    
                    dllCombined['none'].append(
                        common_fields + [
                        
                            'none', # name of the standard
                            1.0,
                            self.aSignalData[dDataLnum[dkey],cnum], # signal
                                                                    # in treat
                            self.aSignalData[dDataLnum[dkey],0], # signal
                                                                 # in control
                            (self.aCvarData[dDataLnum[dkey],cnum] *
                             self.aSignalData[dDataLnum[dkey],cnum]), # SD tr
                            (self.aCvarData[dDataLnum[dkey],0] *
                             self.aSignalData[dDataLnum[dkey],0]), # SD co
                            dPratio['none'], # ratio of phosphorylated
                            dCtrlPratio['none'], # r. of ph. in control
                            self.fold_change(
                                dPratio['none'],
                                dCtrlPratio['none']
                            ) # FC
                        ]
                    )
                    
                    for stdshort, stdname in iteritems(ctrls):
                        
                        dllCombined[stdshort].append(
                            common_fields + [
                                
                                stdshort, # name of the standard
                                self.dStd['Signal'][stdname][cnum], # standard
                                                                    # signal
                                self.daSignalNorm[stdshort][
                                    dDataLnum[dkey],cnum], # signal in treat.
                                self.daSignalNorm[stdshort][
                                    dDataLnum[dkey],0], # signal in control
                                (self.aCvarData[dDataLnum[dkey],cnum] *
                                self.daSignalNorm[stdshort][
                                    dDataLnum[dkey],cnum]), # SD tr
                                (self.aCvarData[dDataLnum[dkey],0] *
                                self.daSignalNorm[stdshort][
                                    dDataLnum[dkey],0]), # SD co
                                dPratio[stdshort], # ratio of phosphorylated
                                dCtrlPratio[stdshort], # r. of ph. in control
                                self.fold_change(
                                    dPratio[stdshort],
                                    dCtrlPratio[stdshort]
                                )# FC
                            ]
                        )
        
        self.dllCombined = dllCombined
        
        # make all to numpy array
        for std in self.dllCombined.keys():
            
            self.dllCombined[std] = np.array(self.dllCombined[std], dtype = np.object)
        
        self.daCombined = self.dllCombined
        del self.dllCombined
        
        # write the table to outfile
        if to_file:
            
            for std, arr in iteritems(self.daCombined):
                
                fname = self.fnCombined % std
                
                sys.stdout.write('\t:: Writing to `%s`.\n' % fname)
                
                self.table_to_file(arr, fname, self.lSingleHdr)
    
    @staticmethod
    def fold_change(val, ctrl):
        """
        Calculates the fold change as {value}/{control} if 
        {control} < {value} and -1 * {control}/{value} otherwise.
        """
        return \
            0.0 if val == ctrl \
            else  val / ctrl if ctrl < val \
            else -1 * ctrl / val
    
    @staticmethod
    def log2_fold_change(fc):
        """
        Log2 transforms fold change value.
        """
        return 1.0 if fc == 0.0 else np.log2(fc) if fc > 0 else np.log2(-1 / fc)
    
    @classmethod
    def table_to_file(cls, arr, fname, hdr = None):
        """
        Writes contents of an array into file with header optionally.
        """
        
        sTable = cls.array_to_string(arr, hdr)
        
        with open(fname, 'w') as fp:
            
            fp.write(sTable)
    
    @classmethod
    def array_to_string(cls, arr, hdr = None):
        """
        Takes a numpy array and returns it as a string of tabular.
        """
        
        sHdr = ''
        
        if hdr is not None:
            
            sHdr = '%s\n' % '\t'.join(hdr)
        
        lls = cls.array_to_lls(arr)
        
        sMain = '\n'.join(map(lambda row: '\t'.join(row), lls))
        
        return '%s%s' % (sHdr, sMain)
    
    @classmethod
    def array_to_lls(cls, arr):
        """
        Converts numpy array to a list of lists of strings.
        """
        return \
            list(
                map(
                    lambda row:
                        list(
                            map(
                                cls.to_string,
                                row
                            )
                        ),
                    arr
                )
            )
    
    @classmethod
    def to_string(cls, val):
        """
        Converts simple values to strings a uniform way.
        """
        return (
                ('%.08f' % val)
                    if type(val) is float
                else ('%u' % val)
                    if type(val) is int
                else str(val)
            )
    
    def fc_table(self, to_file = False):
        """
        Creates tables with fold changes, and exports to files optionally.
        Calculates z-scores, p-values and t-values.
        """
        
        daFcTable       = {}
        daPTopFcTable   = {}
        daUniqueFcTable = {}
        
        for std, tbl in iteritems(self.daCombined):
            
            daFcTable[std] = {}
            
            for treat in self.lDataCols:
                
                if treat[0] == 'Control':
                    continue
                
                daFcTable[std][treat[0]] = []
                
                for it in tbl[np.where(np.logical_and(tbl[:,7] == treat[0],
                                                      tbl[:,10] == 'p'))]:
                    
                    # uniprot, genesymbol, name,
                    # resaa, resnum, label, fold change
                    daFcTable[std][treat[0]].append(
                        [
                            it[0], it[1],
                            it[2], it[5],
                            it[6], it[8],
                            it[9],
                            (
                                self.effect_as_int(it[0], it[5], it[6])
                                if hasattr(self, 'lAllKinases')
                                else 0
                            ),
                            it[-1]
                        ])
                
                daFcTable[std][treat[0]] = np.array(daFcTable[std][treat[0]],
                                                    dtype = np.object)
                
                logfc = \
                    np.array(
                        list(
                            map(
                                self.log2_fold_change,
                                daFcTable[std][treat[0]][:,-1]
                            )
                        ),
                        dtype = np.float64
                    )
                
                sd = np.std(logfc)
                
                zscore = logfc / sd
                
                ndist = sp.stats.norm()
                
                pval = 2 * ndist.cdf(- np.abs(zscore))
                
                df = len(set(daFcTable[std][treat[0]][:,2])) - 1
                
                tdist = sp.stats.t(df)
                
                tval = tdist.ppf(pval)
                
                shape = [daFcTable[std][treat[0]].shape[0], 1]
                
                daFcTable[std][treat[0]] = \
                    np.hstack((
                        daFcTable[std][treat[0]],
                        logfc.reshape(shape),
                        zscore.reshape(shape),
                        pval.reshape(shape),
                        tval.reshape(shape)
                    ))
        
        self.daFcTable = daFcTable
        
        # selecting the psite with lowest p-value for each protein
        
        for std, arrs in iteritems(self.daFcTable):
            
            daUniqueFcTable[std] = {}
            daPTopFcTable[std]   = {}
            
            for tr, arr in iteritems(arrs):
                
                proteins = set(arr[:,0])
                antibodies = set(map(lambda r: (r[2], r[6]), arr))
                
                daPTopFcTable[std][tr] = []
                daUniqueFcTable[std][tr] = []
                
                for protein in proteins:
                    
                    this_protein = arr[np.where(arr[:,0] == protein)]
                    
                    imin = np.argmin(this_protein[:,10]) # the min p-value
                    
                    # avoid duplicates -- only one per antibody
                    #if this_protein[imin, 5] not in antibodies:
                    daPTopFcTable[std][tr].append(this_protein[imin,:])
                    #antibodies.add(this_protein[imin, 5])
                
                for ab in antibodies:
                    
                    this_ab = arr[np.where(np.logical_and(arr[:,2] == ab[0],
                                                          arr[:,6] == ab[1]))]
                    
                    daUniqueFcTable[std][tr].append(this_ab[0,:])
                
                daPTopFcTable[std][tr]   = np.vstack(daPTopFcTable[std][tr])
                daUniqueFcTable[std][tr] = np.vstack(daUniqueFcTable[std][tr])
        
        self.daPTopFcTable   = daPTopFcTable
        self.daUniqueFcTable = daUniqueFcTable
        
        if to_file:
            
            self.fc_table_to_file(unique = True)
            self.fc_table_to_file(unique = False)
    
    def fc_table_to_file(self, unique = True):
        """
        Writes FC tables to files.
        """
        
        daFc = self.daUniqueFcTable if unique else self.daFcTable
        
        for std, arrs in iteritems(daFc):
            
            for tr, arr in iteritems(arrs):
                
                fname = self.fnFcTable if unique else self.fnFcTopTable
                fname = fname % (tr, std)
                
                sys.stdout.write('\t:: Writing to `%s`.\n' % fname)
                
                self.table_to_file(arr, fname, self.lFcTableHdr)
    
    def ptms_lookup(self):
        """
        Looks up the PTMs on array in the databases data.
        Creates list with kinase-substrate interactions
        targeting the substrates on array.
        """
        self.dKinAssaySub = {}
        self.dKinNum = {}
        
        for uniprot, ptms in iteritems(self.dAssaySub):
            
            if uniprot not in self.dKinAssaySub:
                self.dKinAssaySub[uniprot] = []
            
            kss = self.dPhosDb[uniprot] if uniprot in self.dPhosDb else []
            
            for ptm in ptms:
                
                nKin = 0
                for ks in kss:
                    if ptm in ks:
                        self.dKinAssaySub[uniprot].append(ks)
                        nKin += 1
                
                self.dKinNum[(uniprot,
                                ptm.residue.name,
                                ptm.residue.number)] = nKin
        
        sys.stdout.write('\t:: Found %u kinase-substrate interactions for\n'\
            '\t   %s phosphorylation sites.\n'\
            '\t   %.02f kinase for one site in average.\n'\
            '\t   For %u sites no kinase found.\n'\
            '\t   Additional %u sites could not be found in UniProt sequences.\n'\
            '\t   The total number of sites is %u.\n'\
            '\t   %u of these sites should be valid for this organism, and\n'\
            '\t   %u mismatches despite should match for this organism.\n'\
            '\t   Excluding redundant combinations, overall %u sites mismatches.\n' % (
                sum(map(len, self.dKinAssaySub.values())),
                len(list(filter(bool, self.dKinNum.values()))),
                np.mean(list(self.dKinNum.values())),
                len(list(filter(lambda n: n == 0, self.dKinNum.values()))),
                len(self.setSeqMismatch),
                self.aPsiteAnnot.shape[0],
                len(list(filter(lambda p:
                                    self.ncbi_tax_id in \
                                        self.dOrgSpecUniprot[(p[0], p[3], p[4])],
                                    self.aPsiteAnnot))),
                len(self.setSeqMismatchOrg),
                len(self.setSeqMismatchNameOrg)
            ))
    
    def collect_proteins(self):
        """
        Compiles a list of all kinases and substrates.
        """
        self.lAllProteins = set(self.dKinAssaySub.keys()) | \
            set(
                list(
                    itertools.chain(
                        *map(
                            lambda kss:
                                map(
                                    lambda ks:
                                        ks.domain.protein,
                                    kss
                                ),
                            self.dKinAssaySub.values()
                        )
                    )
                )
            )
    
    def load_network(self, pfile = None):
        if not pfile:
            self.pa.load_omnipath()
        else:
            self.pa.init_network(pfile = pfile)
        self.pa.get_directed()
        del self.pa.dgraph.es['ptm']
    
    def get_network(self, keep_also = [], step1 = True,
                    more_steps = None):
        """
        Creates a subnetwork based on certain criteria.
        
        :param list keep_also: List of additional vertex names
                               to include in core.
        :param bool step1: Whether include not only direct
                           links but one step indirect connections.
        :param int more_steps: Include longer indirect connections. This
                               param defines the number of mediator nodes.
        """
        if not hasattr(self, 'whole') or self.whole is None:
            self.whole = self.pa.dgraph
        
        new = copy.deepcopy(self.whole)
        self.pa.graph = new
        self.pa.dgraph = new
        self.pa.update_vname()
        self.pa.update_vindex()
        vids = dict(zip(new.vs['name'], range(new.vcount())))
        keep = list(map(lambda u: vids[u],
                    filter(lambda u: u in vids, self.lAllProteins)))
        keep_also = list(map(lambda u: vids[u],
                         filter(lambda u: u in vids, keep_also)))
        all_to_keep = set(keep) | set(keep_also)
        delete = set(range(new.vcount())) - all_to_keep
        
        keep_step1 = set([])
        for vid in delete:
            if len(set(self.pa.affected_by(vid)._vs) & all_to_keep):
                if len(set(self.pa._affected_by(vid)._vs) & all_to_keep):
                    keep_step1.add(vid)
        
        all_to_keep = all_to_keep | keep_step1
        delete = delete - keep_step1
        
        if more_steps is not None:
            
            n1 = self.pa._neighborhood(all_to_keep, order = more_steps, mode = 'IN')
            n2 = self.pa._neighborhood(all_to_keep, order = more_steps, mode = 'OUT')
            
            longer_paths = set(list(n1)) & set(list(n2))
            
            all_to_keep = all_to_keep | longer_paths
            delete = delete - longer_paths
        
        self.pa.graph.delete_vertices(list(delete))
        self.pa._directed = None
        self.pa.dgraph = self.pa.graph
        self.pa.update_vname()
        self.pa.update_vindex()
        
        sys.stdout.write('\t:: Nodes: %u -> %u, edges: %u -> %u\n'\
            '\t   Removed %u vertices and %u edges\n' % (
                self.whole.vcount(),
                self.pa.graph.vcount(),
                self.whole.ecount(),
                self.pa.graph.ecount(),
                len(delete),
                self.whole.ecount() - self.pa.graph.ecount()
            ))
    
    def sparsen_network(self, perc):
        """
        Removes a portion of the edges based on certain conditions.
        
        """
        
        dens0 = self.pa.graph.density()
        
        if 'ptm' not in self.pa.graph.es.attributes() or \
            max(map(len, self.pa.graph.es['ptm'])) == 0:
            self.pa.load_ptms()
        
        assay_ptms = set(itertools.chain(*self.dKinAssaySub.values()))
        
        # edges between kinases and substrates in the assay
        es_protected = list(filter(lambda e:
                                        len(set(e['ptm']) & assay_ptms),
                                    self.pa.graph.es))
        
        self.pa.graph.es['refxsrc'] = list(
            np.array(list(map(len, self.pa.graph.es['references']))) * \
            np.array(list(map(len, self.pa.graph.es['sources']))))
        
        self.pa.graph.es['sign'] = list(map(lambda e:
                                        e['dirs'].is_stimulation() or \
                                        e['dirs'].is_inhibition(),
                                    self.pa.graph.es))
        
        self.pa.graph.es['nptm'] = list(map(len, self.pa.graph.es['ptm']))
        
        self.pa.graph.es['mindeg'] = list(map(lambda e: min(
                                            self.pa.graph.vs[e.source].degree(),
                                            self.pa.graph.vs[e.target].degree()),
                                            self.pa.graph.es))
        
        self.pa.graph.es['weight'] = list(
            (np.array(self.pa.graph.es['refxsrc']) + 1) * \
            (np.array(self.pa.graph.es['sign'], dtype = np.int) + 1) * \
            (np.array(self.pa.graph.es['nptm']) + 1) / \
            np.array(self.pa.graph.es['mindeg']))
        
        cutoff = np.percentile(np.array(self.pa.graph.es['weight']), perc)
        
        delete = set(list(map(lambda e: e.index,
                        filter(lambda e:
                            e.index in es_protected or \
                            e['weight'] < cutoff,
                        self.pa.graph.es))))
        
        for v in self.pa.graph.vs:
            this_es = list(itertools.chain(
                self.pa.graph.es.select(_source = v.index),
                self.pa.graph.es.select(_target = v.index)
            ))
            
            eids = set(list(map(lambda e: e.index, this_es)))
            
            # if all edges of one node would be removed,
            # we still keep the one with the highest weight
            if len(eids - delete) == 0:
                
                maxweight = max(map(lambda e:
                                        (e['weight'], e.index),
                                    this_es),
                                key = lambda i: i[0])[1]
                
                if maxweight in delete:
                    delete.remove(maxweight)
        
        self.pa.graph.delete_edges(delete)
        self.pa.update_vindex()
        self.pa.update_vname()
        
        sys.stdout.write('\t:: Sparsening network: %u edges'\
            ' removed, %u have been kept.\n'\
            '\t   Density changed from %.04f to %.04f.\n' % \
                (len(delete), self.pa.graph.ecount(),
                 dens0, self.pa.graph.density()))
    
    def collect_assay_psites(self):
        """
        Collects all phosphosites included in the assay.
        Result saved to `setAssayPsites` attribute.
        """
        
        aFc = list(list(self.daFcTable.values())[0].values())[0]
        
        self.setAssayPsites = set(list(map(self.psite_from_fc_row, aFc)))
    
    def collect_participants(self,
                             attr,
                             get_participant = 'get_kinase_uniprot',
                             type_filter = 'is_phosphorylation',
                             regenerate = False):
        """
        Collects all kinases from the kinase-substrate interactions
        loaded from databases. Creates a unique list and saves it to
        attribute `attr`.
        """
        
        if hasattr(self, attr) and not regenerate:
            return None
        
        get_participant = getattr(self, get_participant)
        type_filter     = getattr(self, type_filter)
        
        setattr(self, attr,
            list(
                set(
                    itertools.chain(
                        *map(
                            lambda kss:
                                list(
                                    map(
                                        get_participant,
                                        filter(
                                            type_filter,
                                            kss[1]
                                        )
                                    )
                                ),
                            iteritems(self.dPhosDb)
                        )
                    )
                )
            )
        )
    
    @staticmethod
    def get_kinase_uniprot(dommot):
        """
        Returns the UniProt of the kinase from a
        pypath.intera.DomainMotif object.
        """
        return dommot.domain.protein
    
    @staticmethod
    def is_phosphorylation(dommot):
        """
        Returns `True` if the type of an enzyme-substrate interaction
        is `phosphorylation`.
        """
        return dommot.ptm.typ == 'phosphorylation'
    
    @staticmethod
    def get_substrate_psite(dommot):
        """
        From a `pypath.intera.DomainMotif` object, gets the identity of
        the phosphosite in a 'UniProt_[residue][number]' format.
        """
        return '%s_%s%u' % (dommot.ptm.protein,
                            dommot.ptm.residue.name,
                            dommot.ptm.residue.number)
    
    def kinase_psite_adj(self, regenerate = False):
        """
        Creates and adjacency matrix of kinases and target phosphorylation
        sites, based on database data. The matrix is a `pandas.DataFrame`
        with values of 1.0 if the kinase phosphorylates the site otherwise
        `numpy.nan` values. The matrix is saved into the `dfKinPsite`
        attribute.
        This method also builds a dictionary of kinase-phosphosite
        relationships which will be useful later to query the targets
        of kinases.
        """
        
        if (
            hasattr(self, 'dsetKinPsite') and
            hasattr(self, 'dsetKinAllPsite') and
            not regenerate
        ):
            return None
        
        self.collect_participants('lAllKinases', regenerate = regenerate)
        self.collect_participants('lAllPsites', 'get_substrate_psite',
                                  regenerate = regenerate)
        
        self.dsetKinPsite    = {}
        self.dsetKinAllPsite = {}
        
        nkin = len(self.lAllKinases)
        npst = len(self.lAllPsites)
        
        self.dfKinPsite    = np.empty(shape = [npst, nkin])
        self.dfKinPsite[:] = np.nan
        self.dfKinPsite    = pd.DataFrame(self.dfKinPsite)
        self.dfKinPsite.columns = self.lAllKinases
        self.dfKinPsite.index   = self.lAllPsites
        
        self.collect_assay_psites()
        
        for sub, kss in iteritems(self.dPhosDb):
            
            for ks in kss:
                
                if self.is_phosphorylation(ks):
                    
                    kin   = self.get_kinase_uniprot(ks)
                    psite = self.get_substrate_psite(ks)
                    
                    if kin not in self.dsetKinPsite:
                        self.dsetKinPsite[kin]    = set([])
                        self.dsetKinAllPsite[kin] = set([])
                    
                    self.dfKinPsite.set_value(psite, kin, 1.0)
                    
                    self.dsetKinAllPsite[kin].add(psite)
                    
                    if psite in self.setAssayPsites:
                        self.dsetKinPsite[kin].add(psite)
    
    @staticmethod
    def psite_from_fc_row(row):
        """
        Returns phosphosite string representation from a row of and FC array.
        """
        return '%s_%s%u' % (row[0], row[3], row[4])
    
    def fc_data_frames(self, std = 'none', colname = '30min'):
        """
        Creates a `pandas.DataFrame` to serve as input for `kinact`.
        Data frame contains phosphorylation sites and their fold changes.
        """
        
        self.ddfFcData = {}
        
        for treatment, fctab in iteritems(self.daFcTable[std]):
            
            psites, fc = \
                zip(
                    *map(
                        lambda row:
                            (
                                self.psite_from_fc_row(row), # psite
                                row[9] # fold change
                            ),
                        fctab
                    )
                )
            
            self.ddfFcData[treatment] = pd.DataFrame(np.array(fc),
                                                     index = psites,
                                                     columns = [colname])
    
    def kinact_analysis(self, std = 'none'):
        """
        Sets up and runs kinase activity analysis with `kinact`.
        Results saved to the `dddfKinactResult` attribute.
        """
        
        self.dddfKseaResult      = {}
        self.dddfNetworkinResult = {}
        
        # collecting UniProts of all kinases
        # and identities of all psites
        self.collect_participants('lAllKinases')
        self.collect_participants('lAllPsites', 'get_substrate_psite')
        self.kinase_psite_adj()
        self.fc_data_frames(std = std)
        
        for tr in self.daFcTable[std].keys():
            
            self.dddfKseaResult[tr] = {}
            self.dddfNetworkinResult[tr] = {}
            
            self.dddfKseaResult[tr]['score'], \
            self.dddfKseaResult[tr]['pval'] = \
                                              \
                kinact.ksea.ksea_mean(
                    data_fc = self.ddfFcData[tr]['30min'].dropna(),
                    interactions = self.dfKinPsite,
                    mP = self.ddfFcData[tr]['30min'].values.mean(),
                    delta = self.ddfFcData[tr]['30min'].values.std())
            
            self.dddfNetworkinResult[tr]['score'], \
            self.dddfNetworkinResult[tr]['pval'] = \
                                                   \
                kinact.networkin.weighted_mean(
                    data_fc =self.ddfFcData[tr]['30min'].dropna(),
                    interactions = self.dfKinPsite,
                    mP = self.ddfFcData[tr]['30min'].values.mean(),
                    delta = self.ddfFcData[tr]['30min'].values.std())
    
    def targets_of_kinase(self, kinase_uniprot, on_assay = True, uniprots = False):
        """
        Returns a set of one kinase's targets.
        
        :param str kinase_uniprot: The UniProt ID of the kinase.
        :param bool on_assay: Return the targets only on the assay,
                              or all from the databases.
        :param bool uniprots: Return UniProts, or GeneSymbols otherwise.
        """
        
        dsetKinPsite = self.dsetKinPsite if on_assay else self.dsetKinPsite
        
        setSubUniprots = set(map(lambda ps: ps.split('_')[0],
                                 dsetKinPsite[kinase_uniprot]))
        
        if not uniprots:
            
            return \
                set(
                    itertools.chain(
                        *map(
                            lambda u:
                                self.pa.mapper.map_name(u,
                                                        'uniprot',
                                                        'genesymbol',
                                                        self.ncbi_tax_id),
                            setSubUniprots
                        )
                    )
                )
        
        return setSubUniprots
    
    def psite_kinase_adj(self, regenerate = False):
        """
        Creates a dict between psites and kinases.
        This is for the opposite way lookup as `dsetKinPsite`.
        """
        
        if hasattr(self, 'dsetPsiteKin') and not regenerate:
            return None
        
        self.collect_participants('lAllKinases', regenerate = regenerate)
        self.collect_participants('lAllPsites', 'get_substrate_psite',
                                  regenerate = regenerate)
        
        self.dsetPsiteKin = {}
        
        for kinase, psites in iteritems(self.dsetKinPsite):
            
            for psite in psites:
                
                tPsite = self.get_psite(psite)
                
                if tPsite not in self.dsetPsiteKin:
                    self.dsetPsiteKin[tPsite] = set([])
                
                self.dsetPsiteKin[tPsite].add(kinase)
    
    def get_psite(self, substrate, residue = None, offset = None):
        """
        From a phosphosite string representation or substrate UniProt ID,
        residue name and number separately, returns an uniform tuple
        representation.
        """
        
        if residue is None and offset is None:
            mPsite = self.rePsite.match(substrate)
            if mPsite is None:
                sys.stdout.write('\t:: Could not determine residue and number.\n')
                return None
            substrate, residue, offset = mPsite.groups()
        
        return (substrate, residue, int(offset))
    
    def kinases_of_substrate(self, substrate, residue = None, offset = None):
        """
        Returns the kinases of a substrate.
        """
        
        self.psite_kinase_adj()
        
        psite = self.get_psite(substrate, residue, offset)
        
        return self.dsetPsiteKin[psite] if psite in self.dsetPsiteKin else set([])
    
    def regulatory_sites(self, _reload = False):
        """
        Loads the PhosphoSitePlus regulatory sites data.
        Data about the phosphorylation sites on the assay
        is saved to `ddRegSites`.
        """
        
        if hasattr(self, 'ddRegSites') and not _reload:
            return None
        
        self.kinase_psite_adj()
        
        raw = (
            pypath.dataio.regsites_one_organism(organism = self.ncbi_tax_id)
        )
        
        self.ddRegSites = {}
        
        for psite in self.setAssayPsites:
            
            tPsite = self.get_psite(psite)
            key = (tPsite[1], tPsite[2], 'phosphorylation')
            self.ddRegSites[tPsite] = None
            
            if tPsite[0] in raw:
                
                key = (tPsite[1], tPsite[2], 'phosphorylation')
                
                if key in raw[tPsite[0]]:
                    
                    self.ddRegSites[tPsite] = raw[tPsite[0]][key]
    
    def get_regulation_data(self, key, substrate,
                            residue = None, offset = None):
        """
        Returns a piece of information from the regulatory sites
        data about one phosphorylation site.
        """
        self.regulatory_sites()
        
        defaults = {
            'induces':  set([]),
            'disrupts': set([]),
            'positive': False,
            'negative': False
        }
        
        psite = self.get_psite(substrate, residue, offset)
        
        return (
            self.ddRegSites[psite][key]
            if psite in self.ddRegSites
            and self.ddRegSites[psite] is not None
            else defaults[key]
        )
    
    def induced_by(self, substrate, residue = None, offset = None):
        """
        Returns the UniProt IDs of the interacting protein partners of the
        phosphorylated substrate which the phosphorylation event induces
        the substrate's interaction with.
        """
        
        return self.get_regulation_data('induces', substrate,
                                        residue, offset)
    
    def disrupted_by(self, substrate, residue = None, offset = None):
        """
        Returns the UniProt IDs of the interacting protein partners of the
        phosphorylated substrate which the phosphorylation event disrupts
        the substrate's interaction with.
        """
        
        return self.get_regulation_data('disrupts', substrate,
                                        residue, offset)
    
    def affected_by(self, substrate, residue = None, offset = None):
        """
        Returns the UniProt IDs of the interacting protein partners of the
        phosphorylated substrate which the phosphorylation event affects
        the substrate's interaction with.
        """
        
        return (
            self.get_regulation_data('induces', substrate,
                                        residue, offset) |
            self.get_regulation_data('disrupts', substrate,
                                        residue, offset)
        )
    
    def psite_effect(self, substrate, residue = None, offset = None):
        """
        Returns a tuple of two boolean values.
        The first is True if the phosphorylation event has a
        stimulatory effect. The second is True if the phosphorylation
        has an inhibitory effect.
        """
        return(
            self.get_regulation_data('positive', substrate, residue, offset),
            self.get_regulation_data('negative', substrate, residue, offset)
        )
    
    def effect_as_int(self, substrate, residue = None, offset = None):
        """
        Returns the unambiguous effect as an integer, either -1, 0 or 1.
        """
        return (
            0 +
            int(self.psite_stimulatory_unambiguous(
                substrate, residue, offset)) -
            int(self.psite_inhibitory_unambiguous(
                substrate, residue, offset))
        )
    
    def psite_stimulatory_unambiguous(self, substrate,
                                      residue = None, offset = None):
        """
        Returns True only if the phosphorylation site is stimulatory
        and not inhibitory.
        """
        effect = self.psite_effect(substrate, residue, offset)
        return effect[0] and not effect[1]
    
    def psite_inhibitory_unambiguous(self, substrate,
                                     residue = None, offset = None):
        """
        Returns True only if the phosphorylation site is inhibitory
        and not stimulatory.
        """
        effect = self.psite_effect(substrate, residue, offset)
        return effect[1] and not effect[0]
    
    def psite_stimulatory(self, substrate, residue = None, offset = None):
        """
        Tells if the phosphorylation site has a stimulatory effect.
        """
        return self.get_regulation_data('positive', substrate,
                                        residue, offset)
    
    def psite_inhibitory(self, substrate, residue = None, offset = None):
        """
        Tells if the phosphorylation site has a inhibitory effect.
        """
        return self.get_regulation_data('negative', substrate,
                                        residue, offset)
    
    def kinact_top(self, fname = None, threshold = 0.2):
        """
        Prints a table with top results from `kinact`.
        """
        
        fp = sys.stdout if fname is None else open(fname, 'w')
        
        cwidth = 10
        
        empty = ' ' * (cwidth + 2)
        rule  = '=' * (cwidth + 2)
        
        def cell(content):
            if hasattr(content, 'ljust'):
                return ' %s ' % content.ljust(cwidth)
            if type(content).__name__[:5] == 'float':
                content = '%.05f' % content
            if type(content) is int:
                content = '%u' % content
            return ' %s ' % content.rjust(cwidth)
        
        for meth in ['Ksea', 'Networkin']:
            
            result = getattr(self, 'dddf%sResult' % meth)
            
            numtr  = len(result.keys())
            
            # header
            fp.write('%s\n' % (rule * numtr * 3))
            fp.write('%s\n' % meth)
            fp.write('%s\n' % (rule * numtr * 3))
            fp.write('%s\n' %
                ''.join(
                    list(
                        map(
                            lambda tr:
                                '%s%s%s' % (
                                    cell(tr), cell(empty), cell(empty)
                                ),
                            sorted(result.keys())
                        )
                    )
                )
            )
            fp.write('%s\n' % (rule * numtr * 3))
            fp.write('%s\n' %
                ''.join([
                    '%s%s%s' % (
                        cell('Protein'),
                        cell('Score'),
                        cell('p-value')
                    )
                ] * numtr)
            )
            fp.write('%s\n' % (rule * numtr * 3))
            
            data = {}
            
            for tr in sorted(result.keys()):
                
                res  = result[tr]
                ordr = res['pval'].argsort()
                
                data[tr] = \
                    list(
                        map(
                            lambda i:
                                (
                                    self.pa.mapper.map_name(
                                        res['pval'].index[i],
                                        'uniprot',
                                        'genesymbol',
                                        self.ncbi_tax_id)[0],
                                    res['score'][i],
                                    res['pval'][i]
                                ),
                            filter(
                                lambda i:
                                    res['pval'][i] <= threshold,
                                ordr
                            )
                        )
                    )
            
            nrows = max(map(len, data.values()))
            
            for i in xrange(nrows):
                
                for tr in sorted(result.keys()):
                    
                    dat = data[tr]
                    
                    if i < len(dat):
                        
                        fp.write('%s%s%s' % tuple(map(cell, dat[i])))
                    
                    else:
                        
                        fp.write(empty * 3)
                
                fp.write('\n')
            
            fp.write('%s\n' % (rule * numtr * 3))
        
        if fname is not None:
            
            fp.close()
    
    def fc_diff_table(self):
        """
        Orders proteins by fold change diff.
        """
        
        for attr in ['daFcTable', 'daUniqueFcTable']:
            
            tables = getattr(self, attr)
            
            for std in tables.keys():
                
                tab = tables[std]
                
                ordr = \
                    np.array(
                        list(
                            map(
                                lambda it:
                                    it[0],
                                sorted(
                                    map(
                                        lambda i:
                                            (
                                                i,
                                                max(
                                                    map(
                                                        lambda tr:
                                                            abs(
                                                                tab[tr[0]][i,8]-
                                                                np.sign(
                                                                    tab[tr[0]][i,8]
                                                                ) - (
                                                                tab[tr[1]][i,8] -
                                                                np.sign(
                                                                    tab[tr[1]][i,8]
                                                                )
                                                                )
                                                            ),
                                                        itertools.combinations(
                                                            tab.keys(),
                                                            2
                                                        )
                                                    )
                                                )
                                            ),
                                        xrange(list(tab.values())[0].shape[0])
                                    ),
                                    key = lambda it: it[1],
                                    reverse = True
                                )
                            )
                        )
                    )
                
                for tr in tab.keys():
                    
                    tab[tr] = tab[tr][ordr,:]
        
        self.fc_table_to_file(unique = True)
        self.fc_table_to_file(unique = False)
    
    def fc_top_table(self, unique = True):
        """
        Writes FC values sorted by maximum difference into file.
        """
        
        lHdr = ['uniprot', 'genesymbol', 'name',
                'resaa', 'resnum', 'label',
                'original_resnum', 'effect', 'psite']
        
        self.daFcTop = {}
        
        self.fc_diff_table()
        
        fctab = self.daUniqueFcTable if unique else self.daFcTable
        trs   = sorted(list(fctab.values())[0].keys())
        
        lHdr.extend(trs)
        
        for std, da in iteritems(fctab):
            
            ll = []
            
            for i in xrange(list(da.values())[0].shape[0]):
                
                row = list(list(da.values())[0][i,:8])
                row.append('%s_%s%u' % (row[1], row[3], row[4]))
                
                for tr in trs:
                    
                    row.append(da[tr][i,9])
                
                ll.append(row)
            
            self.daFcTop[std] = np.array(ll, dtype = np.object)
            
            fname = self.fnFcTopCommon % std
            
            sys.stdout.write('\t:: Writing to `%s`.\n' % fname)
            
            self.table_to_file(self.daFcTop[std], fname, lHdr)
    
    def top_fc_venn(self, threshold = 1.0,
                    sign = None,
                    title = '',
                    number_labels = False,
                    figsize = [10, 10],
                    adj = {},
                    treatments = ['BMP8b', 'BMP8b_NE', 'NE'],
                    label_size = None):
        """
        Creates a Venn diagram of the highest FCs by treatments.
        
        :param str,None sign: If None (default) effect signs are not
                              considered. If `positive`, only stimulated
                              if `negative` only inhibited sites used.
        :param bool number_labels: If `True`, number of elements plotted
                                   in each field of the diagram; otherwise
                                   a list of protein names.
        :param str title: The main title of the plot.
        :param float threshold: The threshold for log2 fold change.
                                Fold changes with lower absolute value are
                                not considered.
        :param list figsize: Figure size for matplotlib.
        """
        
        def set_label_text(elements, color, field_id,
                           number = False, _adj = (0.0, 0.0)):
            
            if number:
                text = '%u' % len(elements)
            else:
                text = ', '.join(sorted(elements))
            
            label = plot['venn'].get_label_by_id(field_id)
            
            if label is not None:
                
                label.set_text(text)
                label.set_wrap(True)
                label.set_bbox({
                    'facecolor': 'none',
                    'edgecolor': color,
                    'boxstyle': 'round,pad=1'
                })
                pos = label.get_position()
                label.set_position((pos[0] + _adj[0],
                                    pos[1] + _adj[1]))
                if label_size is not None:
                    label.set_size(label_size)
                
                label_artists.append(label)
        
        def get_adj(setkey):
            return adj[setkey] if setkey in adj else (0.0, 0.0)
        
        treatments = sorted(treatments)
        label_artists = []
        
        if not hasattr(self, 'plotVenn'):
            self.plotVenn = {}
        
        key = ('nosign' if sign is None else sign,
               '%.02f' % threshold)
        self.dsetTopFc = {}
        plot  = {}
        plot['fname'] = self.fnVennPlot % key
        
        for tr in treatments:
            
            arr = self.daUniqueFcTable['none'][tr]
            # we either select the 
            if sign == 'positive':
                arr = arr[np.where(arr[:,7] ==  1)[0], :]
            elif sign == 'negative':
                arr = arr[np.where(arr[:,7] == -1)[0], :]
            
            if sign is None:
                values = np.abs(arr[:,9])
            else:
                values = arr[:,9] * arr[:,7]
            
            self.dsetTopFc[tr] = (
                set(arr[np.where(values >= threshold)[0], 1])
            )
        
        plot['pdf'] = mpl.backends.backend_pdf.PdfPages(plot['fname'])
        plot['fig'] = mpl.figure.Figure(figsize = figsize)
        plot['cvs'] = mpl.backends.backend_pdf.FigureCanvasPdf(plot['fig'])
        plot['ax']  = plot['fig'].add_subplot(1, 1, 1)
        
        plot['venn'] = mplvenn.venn3(
            subsets = list(map(lambda tr: self.dsetTopFc[tr],
                               sorted(self.dsetTopFc.keys()))),
            set_labels = sorted(self.dsetTopFc.keys()),
            ax = plot['ax']
        )
        
        # only in one set
        for tr, field_id, color in zip(
            treatments,
            ['100', '010', '001'],
            ['#CC5555', '#55AA55', '#5555CC']):
            
            elements  = self.dsetTopFc[tr]
            othertr   = set(treatments) - set([tr])
            otherelem = set().union(*map(lambda otr: self.dsetTopFc[otr],
                                         othertr))
            elements = elements - otherelem
            set_label_text(elements, color, field_id,
                        number_labels, _adj = get_adj(tr))
            plot[tr] = elements
        
        # intersections of 2 sets
        for pair, field_id, color in zip(
                itertools.combinations(treatments, 2),
                ['110', '101', '011'],
                ['#C0A055', '#C044C0', '#44A0C0']
            ):
            
            other    = list(set(treatments) - set(pair))[0]
            elements = set().intersection(*map(lambda tr: self.dsetTopFc[tr], pair))
            elements = elements - self.dsetTopFc[other]
            
            set_label_text(elements, color, field_id,
                           number_labels, _adj = get_adj('%s:%s' % pair))
            
            plot['%s:%s' % pair] = elements
        
        # intersection of all the 3
        elements = set().intersection(*list(self.dsetTopFc.values()))
        set_label_text(elements, '#A088A0', '111',
                       number_labels, _adj = get_adj(':'.join(treatments)))
        plot[':'.join(treatments)] = elements
        
        plot['ax'].set_title(title)
        plot['fig'].tight_layout(pad = 1.0, rect = (0.1, 0.1, 0.8, 0.8))
        #plot['fig'].subplots_adjust(bottom = 0.2, top = 0.8, right = 0.8, left = 0.2)
        plot['cvs'].draw()
        plot['cvs'].print_figure(plot['pdf'])
                                 #bbox_inches = 'tight',
                                 #bbox_extra_artists = label_artists)
        plot['pdf'].close()
        plot['fig'].clf()
        
        self.plotVenn[key] = plot
    
    def antibody_id_to_name(self, aid, unique = True):
        """
        For one antibody ID returns the protein names and residues
        as a human readable name.
        If ``unique``, it uses a counter and appends an additional
        number in parentheses to make a series of labels unique.
        The counter is under attribute ``cntrUniqueLabels``.
        """
        regsymbol = re.compile(r'([A-Z0-9]*?)([0-9]*[A-Z]?$)')
        repost = re.compile(r'([0-9]*)([A-Z]*)')
        
        if type(aid) is int:  aid = [aid]
        if type(aid) is list: aid = np.array(aid)
        
        def shorten_protein_names(proteins):
            gsymbols = \
                list(
                    map(
                        lambda gs:
                            regsymbol.match(gs).groups(),
                        proteins
                    )
                )
            
            stems = collections.Counter(map(lambda gs: gs[0], gsymbols))
            
            short = (
                '/'.join(
                    sorted(
                        map(
                            lambda gss:
                                '%s%s%s%s' % (
                                    gss,
                                    '[' if stems[gss] > 1 else '',
                                    ','.join(
                                        sorted(
                                            list(
                                                map(
                                                    lambda gs:
                                                        gs[1],
                                                    filter(
                                                        lambda gs:
                                                            gs[0] == gss,
                                                        gsymbols
                                                    )
                                                )
                                            ),
                                            key = lambda post:
                                                (
                                                    int(repost.match(post).group(1))
                                                    if repost.match(post).group(1).isdigit() else 0,
                                                    repost.match(post).group(2)
                                                )
                                        )
                                    ),
                                    ']' if stems[gss] > 1 else ''
                                ),
                            sorted(stems.keys())
                        )
                    )
                )
            )
            
            return 'Histone H3' if short[:2] == 'H3' else short
        
        proteins, psites = \
            list(
                zip(
                    *map(
                        lambda ps:
                            (
                                self.pa.mapper.map_name(ps[0], 'uniprot', 'genesymbol',
                                                        ncbi_tax_id = self.ncbi_tax_id),
                                (ps[3], ps[4])
                            ),
                        self.aPsiteAnnot[np.where(np.in1d(self.aPsiteAnnot[:,5], aid))]
                    )
                )
            )
        
        label = (
            '%s-%s' % (
                shorten_protein_names(set(itertools.chain(*proteins))),
                '/'.join(
                    map(
                        lambda ps:
                            '%s%u' % ps,
                        sorted(
                            set(psites),
                            key = lambda ps: ps[1]
                        )
                    )
                )
            )
        )
        
        if unique:
            # labels must be unique
            self.cntrUniqueLabels.update([label])
            
            if self.cntrUniqueLabels[label] > 1:
                label = '%s(%u)' % (label, self.cntrUniqueLabels[label])
        
        return label
    
    ### Direct functional annotation
    #
    # Categories and parent terms:
    #
    # -- angiogenesis: GO:0001525
    # -- neurogenesis: GO:0022008
    #                  GO:0038179 (neurotrophin signaling pathway)
    # -- inflammation: GO:0006954
    # -- lipid metabolism: GO:0006629 or
    #                      GO:0044255 (cellular lipid metabolic process) or
    #                      GO:0045444 (fat cell differentiation)
    # -- survival, cell cycle: GO:0010941 (regulation of cell death) or
    #                          GO:0007049 (cell cycle)
    #
    
    def load_go(self, _reload = False):
        """
        Loads the GO biological process annotations
        if those haven't been loaded or reload forced.
        Also loads the ontology graph as a
        ``pysemsim.GOTrees`` object. This is for lookup
        of ancestors.
        """
        
        if (
            hasattr(self, 'dsetGOBP') and
            hasattr(self, 'dsGONames') and
            hasattr(self, 'GOOntology') and
            not _reload
        ):
            return None
        
        go = pypath.dataio.get_go_quick(organism = self.ncbi_tax_id)
        
        self.dsetGOBP    = go['terms']['P']
        self.dsGONames = go['names']
        
        urlObo = pypath.urls.urls['go']['url']
        c = pypath.curl.Curl(urlObo, silent = False, large = True)
        fnObo = c.fileobj.name
        c.close()
        del c
        
        self.GOOntology     = pysemsim.GOTrees(fnObo)
        self.dsetGOBPAncestors = (
            dict(
                map(
                    lambda protein: (
                        protein,
                        self.terms_with_ancestors(protein)
                    ),
                    self.dsetGOBP.keys()
                )
            )
        )
    
    def read_functional_categories(self, reread = False):
        """
        Reads a set of functional categories defined by GO terms.
        """
        
        if hasattr(self, 'dsetFuncCat') and not reread:
            return None
        
        self.dsetFuncCat = {}
        
        with open(self.fnFuncCat, 'r') as fp:
            
            for l in fp:
                
                l = l.split('\t')
                
                if l[0] not in self.dsetFuncCat:
                    self.dsetFuncCat[l[0]] = set([])
                
                self.dsetFuncCat[l[0]].add(l[1].split()[0])
    
    def terms_with_ancestors(self, protein):
        """
        Returns all GO terms of one ``protein`` including all their
        ancestors.
        """
        return (
            set(
                itertools.chain(
                    *map(
                        lambda term:
                            self.GOOntology.get_ancestors(term),
                        self.dsetGOBP[protein]
                    )
                )
            ) | self.dsetGOBP[protein]
        )
    
    def terms_of_category(self, protein, cat):
        """
        Returns the terms in one functional category ``cat``
        one ``protein`` has been annotated with.
        """
        
        return self.dsetGOBPAncestors[protein] & self.dsetFuncCat[cat]
    
    def has_terms_of_category(self, protein, cat):
        """
        Returns if one ``protein`` has been annotated with any of
        the terms in one functional category ``cat``.
        """
        
        return bool(self.terms_of_category(protein, cat))
    
    def ratio_having_terms_of_category(self, proteins, cat):
        """
        Returns the ratio of the proteins in list ``proteins``
        that have been annotated with any terms in category ``cat``.
        """
        
        return (
            sum(map(lambda p: self.has_terms_of_category(p, cat), proteins)) /
            float(len(proteins))
        ) if len(proteins) else 0.0
    
    def functional_array(self):
        """
        Compiles an array of functional annotations.
        """
        
        self.read_functional_categories()
        
        arr    = list(self.daUniqueFcTable['none'].values())[0]
        cats   = sorted(self.dsetFuncCat.keys())
        llFunc = []
        
        for row in arr:
            
            psite   = (row[0], row[3], row[4])
            protein = row[0]
            
            effect = self.effect_as_int(*psite)
            
            self_func = (
                list(
                    map(
                        lambda cat:
                            int(self.has_terms_of_category(protein, cat)),
                        cats
                    )
                )
            )
            
            kinases_func = (
                list(
                    map(
                        lambda cat:
                            self.ratio_having_terms_of_category(
                                self.kinases_of_substrate(*psite),
                                cat
                            ),
                        cats
                    )
                )
            )
            
            regulated_func = (
                list(
                    map(
                        lambda cat:
                            self.ratio_having_terms_of_category(
                                self.affected_by(*psite),
                                cat
                            ),
                        cats
                    )
                )
            )
            
            llFunc.append(
                list(
                    itertools.chain(
                        psite,
                        [row[5], effect],
                        self_func,
                        kinases_func,
                        regulated_func
                    )
                )
            )
            
        
        lHdrFunc = (
            list(
                itertools.chain(
                    ['uniprot', 'resaa', 'resnum', 'label', 'effect'],
                    map(lambda cat: '%s;self' % cat, cats),
                    map(lambda cat: '%s;kinases' % cat, cats),
                    map(lambda cat: '%s;regulated' % cat, cats),
                )
            )
        )
        
        self.aFunc = np.array(llFunc, dtype = np.object)
        self.lHdrFunc = lHdrFunc
    
    def tidy_functional_array(self):
        """
        Creates a tidy array from the functional data and exports it to file.
        """
        
        lHdr = ['label', 'func', 'category', 'value']
        
        func_of = ['self', 'kinases', 'regulated']
        cats    = sorted(self.dsetFuncCat.keys())
        result  = []
        
        for icat, cat in enumerate(cats):
            
            for ifun, fun in enumerate(func_of):
                
                col = 5 + len(cats) * ifun + icat
                
                result.append(
                    np.hstack([
                        self.aFunc[:,3].reshape(self.aFunc.shape[0], 1),
                        np.array([[cat, fun]] * self.aFunc.shape[0]),
                        self.aFunc[:,col].reshape(
                            self.aFunc.shape[0], 1).astype(np.float)
                    ])
                )
        
        self.aTidyFunc = np.vstack(result)
        
        self.table_to_file(self.aTidyFunc, self.fnFuncTidy, lHdr)
    
    def fc_list(self, psites, treatment, signs = False,
                category = None, std = 'none'):
        """
        Returns fold change values in one treatment
        for a list of phosphosites, optionally with
        signes corrected according to the effect of
        each phosphosite.
        """
        
        tbl = self.daUniqueFcTable[std][treatment]
        
        link_methods = {
            'kinases':   self.kinases_of_substrate,
            'regulated': self.affected_by
        }
        
        if category is not None and category != 'self':
            tbl = tbl[
                np.where(
                    np.array(
                        list(
                            map(
                                lambda r:
                                    bool(len(
                                        link_methods[category](
                                            r[0], r[3], r[4]))),
                                tbl
                            )
                        )
                    )
                )]
        
        psites = np.array(list(psites))
        
        if signs:
            have_sign = tbl[np.where(
                np.in1d(tbl[np.where(tbl[:,7] != 0)][:,5], psites, assume_unique = True)
            )]
            
            return have_sign[:,9] * have_sign[:,7]
        
        else:
            return np.abs(tbl[np.where(
                np.in1d(tbl[:,5], psites, assume_unique = True)
            )][:,9])
    
    def fc_of_annotated_with(self, annotation, treatment,
                             signs = False, category = None,
                             std = 'none'):
        """
        Returns the FC values of from phosphosites in one treatment annotated
        with a certain functional annotation.
        """
        if category is not None:
            arr = self.aTidyFunc[np.where(self.aTidyFunc[:,2] == category)]
        else:
            arr = self.aTidyFunc
        
        setAllPsites = set(self.aTidyFunc[:,0])
        
        annotated = (
            set(
                arr[
                    np.where(
                        np.logical_and(
                            arr[:,1] == annotation,
                            arr[:,3] > 0
                        )
                    )
                ][:,0]
            )
        )
        
        return (
            self.fc_list(annotated, treatment,
                         signs = signs, category = category, std = std),
            self.fc_list(setAllPsites - annotated, treatment,
                         signs = signs, category = category, std = std)
        )
    
    def functional_compare_fc(self, annotation, treatment,
                              silent = False, return_result = False,
                              signs = False,
                              category = None,
                              std = 'none'):
        """
        Performs a Mann-Whitney U-test between fold changes of phosphosites
        annotated or not with certain functionality.
        """
        
        aPos, aNeg = self.fc_of_annotated_with(annotation,
                                               treatment,
                                               signs = signs,
                                               category = category,
                                               std = std)
        
        fUval, fPval = stats.mannwhitneyu(aPos, aNeg,
                                          alternative = 'two-sided'
                                          )
        
        bGreater = aPos.mean() > aNeg.mean()
        
        if not silent:
            sys.stdout.write('\t[%s] In case of %s treatment, '\
                             'looking at annotations of `%s`,\n'\
                             '\t    the fold changes of phosphosites '\
                             'involved in `%s`\n'\
                             '\t    are %s %s; Mann-Whitney p = %.08f; n1 = %u, n2 = %u\n' % (
                                'S' if fPval <= 0.05 else 'N',
                                treatment,
                                'all' if category is None else category,
                                annotation,
                                ('significantly'
                                 if fPval <= 0.05
                                 else 'not significantly'),
                                'greater' if bGreater else 'lower',
                                fPval, len(aPos), len(aNeg)
                            ))
        
        if return_result:
            return fUval, fPval, bGreater
    
    def functional_compare_fcs(self, signs = False, std = 'none'):
        """
        Compares differences in fold changes across all functional annotations
        and all treatments.
        """
        
        for category in [None, 'self', 'kinases', 'regulated']:
            
            for sAnnot in sorted(set(self.aTidyFunc[:,1])):
                
                for sTreat in sorted(self.daUniqueFcTable[std].keys()):
                    
                    self.functional_compare_fc(sAnnot, sTreat,
                                            signs = signs,
                                            category = category,
                                            std = std)
    
